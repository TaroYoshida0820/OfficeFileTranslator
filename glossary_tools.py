"""
glossary_tools.py

用語集(AEC_翻訳用語集.xlsx)を読み込み、
①識別子(信号名・REQ-ID・SPEC-ID等)をAPI呼び出し前にマスキングして保護する
②本文に含まれる一般用語集エントリを抽出し、プロンプトへのヒントとして使う
の2つの機能を提供する。

識別子マスキングの考え方:
    "Scent_Intensity_Level は REQ-016 に対応する" のようなテキストをそのままAPIに渡すと、
    翻訳エンジンが信号名やIDを勝手に訳したり表記を変えたりするリスクがある。
    そこで送信前に "Scent_Intensity_Level" -> "§ID0§" のような一時トークンに置き換え、
    APIには "§ID0§ は §ID1§ に対応する" という形で渡す。
    翻訳後の文中に残った §ID0§ 等のトークンを、応答受領後に元の文字列へ復元する。
    これにより翻訳エンジンの言語能力に関係なく、識別子は機械的に100%保護される。
"""

import re
import openpyxl

# 信号ID・要求ID・仕様書IDのパターン(正規表現)。
# 用語集の "02_識別子_翻訳禁止" シートに列挙されている命名規則に対応。
# 信号ID・要求ID・仕様書IDのパターン(正規表現)。
# 用語集の "02_識別子_翻訳禁止" シートに列挙されている命名規則に対応。
# 注意: \b (単語境界) はPythonのreではUnicode文字も"word文字"とみなすため、
# 日本語の直後(例: "REQ-016に")では境界が成立せず検出漏れする。
# そのためASCII英数字・アンダースコアのみを対象にした否定先読み/後読みで代用する。
_ASCII_WORD = r'[A-Za-z0-9_]'
IDENTIFIER_PATTERNS = [
    re.compile(rf'(?<!{_ASCII_WORD})SIG-(?:RX|TX)-\d+(?!{_ASCII_WORD})'),
    re.compile(rf'(?<!{_ASCII_WORD})REQ-\d+(?!{_ASCII_WORD})'),
    re.compile(rf'(?<!{_ASCII_WORD})SPEC-[A-Z]+-\d+(?!{_ASCII_WORD})'),
    # Scent_Intensity_Level のような「英単語_英単語」形式の信号名(先頭大文字・アンダースコア区切り)
    re.compile(rf'(?<!{_ASCII_WORD})[A-Z][A-Za-z]*(?:_[A-Za-z0-9]+){{1,}}(?!{_ASCII_WORD})'),
]


def load_glossary(xlsx_path):
    """
    AEC_翻訳用語集.xlsx を読み込み、(一般用語リスト, 識別子リテラル集合) を返す。
    一般用語リスト: [{"ja": "...", "en": "...", "note": "..."}]
    識別子リテラル集合: 正規表現パターンに加え、用語集シート2に明示されている個別信号名など
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    terms = []
    if "01_翻訳用語集" in wb.sheetnames:
        ws = wb["01_翻訳用語集"]
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            ja = row_dict.get("日本語")
            en = row_dict.get("英訳")
            if ja and en:
                terms.append({
                    "ja": str(ja).strip(),
                    "en": str(en).strip(),
                    "note": str(row_dict.get("使用文脈・備考") or "").strip(),
                    "ng": str(row_dict.get("NG訳例(誤訳注意)") or "").strip(),
                })

    identifier_literals = set()
    if "02_識別子_翻訳禁止" in wb.sheetnames:
        ws2 = wb["02_識別子_翻訳禁止"]
        headers2 = [c.value for c in ws2[1]]
        for row in ws2.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers2, row))
            ident = row_dict.get("識別子/パターン")
            if ident and "/" not in str(ident) and "xxx" not in str(ident):
                identifier_literals.add(str(ident).strip())

    return terms, identifier_literals


def mask_identifiers(text, extra_literals=None):
    """
    テキスト中の識別子を一時トークンに置換する。
    戻り値: (マスク済みテキスト, {トークン: 元の文字列})
    """
    mapping = {}
    masked = text
    counter = [0]

    def _replace(m):
        token = f"§ID{counter[0]}§"
        mapping[token] = m.group(0)
        counter[0] += 1
        return token

    for pattern in IDENTIFIER_PATTERNS:
        masked = pattern.sub(_replace, masked)

    if extra_literals:
        for literal in sorted(extra_literals, key=len, reverse=True):
            if literal and literal in masked:
                token = f"§ID{counter[0]}§"
                mapping[token] = literal
                counter[0] += 1
                masked = masked.replace(literal, token)

    return masked, mapping


def unmask_identifiers(text, mapping):
    """mask_identifiers()で作ったmappingを使い、トークンを元の識別子に復元する。"""
    for token, original in mapping.items():
        text = text.replace(token, original)
    return text


def find_relevant_terms(text, terms, max_terms=8):
    """本文に含まれていそうな用語集エントリだけを抽出する(軽量版・文字列一致ベース)。
    本格的なベクトルRAGに差し替える場合はこの関数だけ入れ替えればよい設計。"""
    hits = [t for t in terms if t["ja"] and t["ja"] in text]
    return hits[:max_terms]
